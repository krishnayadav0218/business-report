"""
conftest.py
Shared fixtures for the test suite. Pytest auto-discovers this file --
no need to import it manually in test_*.py files.
"""

import os
import sys
import tempfile

import pandas as pd
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


@pytest.fixture
def sample_csv_path(tmp_path):
    """A clean, well-formed CSV -- the easy case."""
    data = {
        "Date": ["01/06/2026", "02/06/2026", "03/06/2026", "04/06/2026"],
        "Region": ["West", "North", "East", "West"],
        "Salesperson": ["Rohit", "Priya", "Arjun", "Rohit"],
        "Product": ["Widget A", "Widget B", "Widget C", "Widget A"],
        "Target": [50000, 60000, 45000, 55000],
        "Collection": [48000, 61200, 42000, 53000],
        "Orders": [12, 15, 10, 13],
        "Cost": [30000, 35000, 28000, 32000],
    }
    path = tmp_path / "sample.csv"
    pd.DataFrame(data).to_csv(path, index=False)
    return str(path)


@pytest.fixture
def sample_df(sample_csv_path):
    """A cleaned DataFrame ready for KPI/insight computation."""
    from src import ingest, process
    raw_df = ingest.from_any_file(sample_csv_path)
    return process.clean_data(raw_df)


@pytest.fixture
def messy_excel_path(tmp_path):
    """Excel with a title row + blank row before the real header -- the messy real-world case."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["ABC Company - Monthly Report"])
    ws.append([])
    ws.append(["Sale Date", "Branch", "Executive Name", "Item", "Monthly Target", "Amount Received", "No. of Orders", "Expense"])
    ws.append(["01-06-2026", "Mumbai", "Rahul", "Plan A", 50000, 48000, 12, 30000])
    ws.append(["02-06-2026", "Delhi", "Sneha", "Plan B", 60000, 61200, 15, 35000])
    path = tmp_path / "messy.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def multisheet_excel_path(tmp_path):
    """Excel where sheet 1 is a cover page and the real data is on sheet 2."""
    from openpyxl import Workbook
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append(["Company Confidential"])

    data_sheet = wb.create_sheet("Data")
    data_sheet.append(["Date", "Region", "Collection", "Target"])
    data_sheet.append(["01-06-2026", "West", 48000, 50000])
    data_sheet.append(["02-06-2026", "North", 61200, 60000])
    path = tmp_path / "multisheet.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def app_client(monkeypatch, tmp_path):
    """A Flask test client with an isolated output directory and no app password."""
    monkeypatch.chdir(tmp_path)
    os.makedirs("output", exist_ok=True)
    os.makedirs("data", exist_ok=True)
    monkeypatch.delenv("APP_PASSWORD", raising=False)

    # Reimport app fresh so it picks up the temp working directory
    import importlib
    import app as app_module
    importlib.reload(app_module)
    return app_module.app.test_client()
